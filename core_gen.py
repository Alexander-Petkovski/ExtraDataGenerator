"""
ExtraDataGenerator — core_gen.py
=================================
Generates intentionally messy CSV / Excel test data for ExtraDataCleaner.

Every issue injected maps 1-to-1 to a cleaning operation in ExtraDataCleaner,
so generated files can be used as direct regression tests.
"""

from __future__ import annotations

import io
import random
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd

# ── static data pools ─────────────────────────────────────────────────────────

_FIRST_NAMES = [
    "Alice", "Bob", "Charlie", "Diana", "Eve", "Frank", "Grace", "Henry",
    "Iris", "Jack", "Karen", "Liam", "Maria", "Noah", "Olivia", "Paul",
    "Quinn", "Rachel", "Sam", "Tina", "Uma", "Victor", "Wendy", "Xavier",
    "Yara", "Zach", "Amy", "Brian", "Chloe", "David", "Emma", "Fred",
    "Gina", "Harry", "Ivy", "James", "Kate", "Leo", "Mia", "Nathan",
    "Oprah", "Peter", "Queenie", "Ross", "Sara", "Tom", "Ursula", "Vince",
    "Wanda", "Xander",
]

_LAST_NAMES = [
    "Smith", "Jones", "Williams", "Brown", "Taylor", "Davies", "Evans",
    "Wilson", "Thomas", "Roberts", "Johnson", "Lewis", "Walker", "Robinson",
    "Wood", "Thompson", "White", "Watson", "Jackson", "Wright", "Green",
    "Harris", "Cooper", "King", "Lee", "Martin", "Clarke", "James",
    "Morgan", "Hughes", "Edwards", "Hill", "Moore", "Clark", "Harrison",
    "Scott", "Young", "Morris", "Hall", "Ward", "Turner", "Carter",
    "Phillips", "Mitchell", "Patel", "Adams", "Campbell", "Anderson",
    "Allen", "Cook",
]

_EMAIL_DOMAINS = [
    "gmail.com", "yahoo.com", "outlook.com", "hotmail.com",
    "company.com", "example.org", "proton.me", "icloud.com",
]

_CATEGORIES = [
    "Electronics", "Clothing", "Food", "Books", "Furniture",
    "Toys", "Sports", "Beauty", "Automotive", "Garden",
]

_STATUS_VALUES = ["Active", "Inactive", "Pending", "Cancelled", "Completed"]

_DEPARTMENTS = [
    "Sales", "Marketing", "Engineering", "HR", "Finance",
    "Operations", "Legal", "Support",
]

_NULL_VARIANTS = [
    "N/A", "n/a", "NA", "na", "NULL", "null", "None", "none",
    "missing", "MISSING", "unknown", "UNKNOWN", "-", "--", "---",
    ".", "#N/A", "undefined", "not available",
]

_BOOL_TRUE  = ["True", "true", "TRUE", "Yes", "yes", "YES", "Y", "y", "1", "On", "on"]
_BOOL_FALSE = ["False", "false", "FALSE", "No", "no", "NO", "N", "n", "0", "Off", "off"]

_EXCEL_ERRORS = ["#VALUE!", "#REF!", "#DIV/0!", "#NAME?", "#NUM!", "#NULL!", "#N/A"]

_CURRENCY_SYMBOLS = ["$", "£", "€", "¥"]

_NOTES_POOL = [
    "Excellent", "Good", "Average", "Needs review", "Pending approval",
    "Approved", "Rejected", "Under investigation", "Follow up required",
    "Completed", "See notes", "Contact customer", "Verify data",
    "Priority item", "On hold", "Escalated", "In progress",
]

# ── column type registry ───────────────────────────────────────────────────────

COLUMN_TYPES: dict[str, str] = {
    "first_name":  "First Name",
    "last_name":   "Last Name",
    "email":       "Email",
    "phone":       "Phone",
    "date":        "Date",
    "integer":     "Count",
    "float":       "Value",
    "currency":    "Amount",
    "percentage":  "Score",
    "boolean":     "Active",
    "category":    "Category",
    "status":      "Status",
    "text":        "Notes",
}

# ── issue registry (key → human label) ────────────────────────────────────────

ISSUE_LABELS: dict[str, str] = {
    "bom":                "BOM prefix on CSV file",
    "title_row":          "Title / junk row before header",
    "blank_rows":         "Blank rows interspersed (~5%)",
    "blank_cols":         "Fully-empty columns added",
    "null_variants":      "Null variants  (N/A, none, NULL…)",
    "eu_decimal":         "EU decimal format  (1.234,56)",
    "currency_symbols":   "Currency symbols  ($, £, €, ¥)",
    "accounting_neg":     "Accounting negatives  (500.00)",
    "percentages":        "Percentage suffix  (75.0%)",
    "excel_errors":       "Excel error strings  (#VALUE! …)",
    "mixed_booleans":     "Mixed boolean formats  (yes/Y/1…)",
    "extra_whitespace":   "Extra whitespace in cells",
    "mixed_case_headers": "Mixed-case column headers",
    "numbers_as_text":    "Numbers stored as text strings",
    "duplicate_col":      "Duplicate column header",
    "merged_cells":       "Merged cells  (Excel only)",
    "hidden_rows_cols":   "Hidden rows / columns  (Excel only)",
}


# ── generator ─────────────────────────────────────────────────────────────────

class DataGenerator:
    """
    Generates reproducibly messy test DataFrames.

    All randomness is seeded so the same seed + config always
    produces identical output — essential for regression testing.
    """

    def __init__(self, seed: int = 42):
        self._seed    = seed
        self._rng     = np.random.default_rng(seed)
        self._py_rng  = random.Random(seed)
        self._report: list[str] = []

    # ── public API ────────────────────────────────────────────────────────────

    def get_report(self) -> str:
        return "\n".join(self._report)

    def generate(
        self,
        rows:       int,
        col_types:  list[str],
        issues:     dict[str, bool],
        fmt:        str,            # "csv" | "xlsx" | "both"
        out_path:   Path,
        filename:   str = "test_data",
    ) -> dict[str, Path | None]:
        """
        Generate messy test data and write to disk.

        Returns {"csv": Path|None, "xlsx": Path|None}.
        """
        self._report = []
        self._rng    = np.random.default_rng(self._seed)
        self._py_rng = random.Random(self._seed)

        # 1. Build clean base DataFrame
        df = self._build_clean(rows, col_types)
        self._report.append(
            f"Base data: {rows} rows × {len(df.columns)} columns"
            f"  (seed={self._seed})"
        )

        # 2. Inject issues
        df = self._inject_issues(df, issues, col_types)

        # 3. Write
        out_path.mkdir(parents=True, exist_ok=True)
        title = f"Exported Records — Test Dataset ({rows} rows)"
        result: dict[str, Path | None] = {"csv": None, "xlsx": None}

        if fmt in ("csv", "both"):
            p = out_path / f"{filename}.csv"
            self._write_csv(
                df, p,
                bom=issues.get("bom", False),
                title=title if issues.get("title_row", False) else None,
            )
            result["csv"] = p
            self._report.append(f"\n  ✓  CSV  →  {p}")
            if issues.get("bom", False):
                self._report.append("       BOM (\\ufeff) prepended.")
            if issues.get("title_row", False):
                self._report.append(f'       Title row prepended: "{title}"')

        if fmt in ("xlsx", "both"):
            p = out_path / f"{filename}.xlsx"
            self._write_xlsx(
                df, p,
                merged=issues.get("merged_cells", False),
                hidden=issues.get("hidden_rows_cols", False),
                title=title if issues.get("title_row", False) else None,
            )
            result["xlsx"] = p
            self._report.append(f"  ✓  Excel  →  {p}")
            if issues.get("title_row", False):
                self._report.append(f'       Title row prepended: "{title}"')

        return result

    # ── clean data builder ────────────────────────────────────────────────────

    def _build_clean(self, rows: int, col_types: list[str]) -> pd.DataFrame:
        fn_pool = self._py_rng.choices(_FIRST_NAMES, k=rows)
        ln_pool = self._py_rng.choices(_LAST_NAMES,  k=rows)

        data: dict[str, list] = {}
        for ctype in col_types:
            col = COLUMN_TYPES.get(ctype, ctype)
            if ctype == "first_name":
                data[col] = fn_pool[:]
            elif ctype == "last_name":
                data[col] = ln_pool[:]
            elif ctype == "email":
                doms = self._py_rng.choices(_EMAIL_DOMAINS, k=rows)
                data[col] = [
                    f"{f.lower()}.{l.lower()}@{d}"
                    for f, l, d in zip(fn_pool, ln_pool, doms)
                ]
            elif ctype == "phone":
                data[col] = [
                    f"+1-{self._py_rng.randint(200,999)}"
                    f"-{self._py_rng.randint(100,999)}"
                    f"-{self._py_rng.randint(1000,9999)}"
                    for _ in range(rows)
                ]
            elif ctype == "date":
                base    = pd.Timestamp("2020-01-01")
                offsets = self._rng.integers(0, 365 * 4, rows)
                data[col] = [
                    (base + pd.Timedelta(days=int(d))).strftime("%Y-%m-%d")
                    for d in offsets
                ]
            elif ctype == "integer":
                data[col] = self._rng.integers(1, 10_000, rows).tolist()
            elif ctype == "float":
                data[col] = np.round(
                    self._rng.uniform(-500.0, 500.0, rows), 2
                ).tolist()
            elif ctype == "currency":
                data[col] = np.round(
                    self._rng.uniform(0.01, 9_999.99, rows), 2
                ).tolist()
            elif ctype == "percentage":
                data[col] = np.round(
                    self._rng.uniform(0.0, 100.0, rows), 1
                ).tolist()
            elif ctype == "boolean":
                data[col] = [
                    bool(v) for v in self._rng.choice([True, False], rows)
                ]
            elif ctype == "category":
                data[col] = self._py_rng.choices(_CATEGORIES, k=rows)
            elif ctype == "status":
                data[col] = self._py_rng.choices(_STATUS_VALUES, k=rows)
            elif ctype == "text":
                data[col] = self._py_rng.choices(_NOTES_POOL, k=rows)

        return pd.DataFrame(data)

    # ── issue injection ───────────────────────────────────────────────────────

    def _inject_issues(
        self,
        df:        pd.DataFrame,
        issues:    dict[str, bool],
        col_types: list[str],
    ) -> pd.DataFrame:
        df = df.copy()

        def col_for(ctype: str) -> str | None:
            """Return actual column name (handles renames)."""
            return self._find_col(df, COLUMN_TYPES.get(ctype, ctype))

        # ── 1. Currency symbols ───────────────────────────────────────────────
        if issues.get("currency_symbols", False):
            c = col_for("currency")
            if c:
                sym = self._py_rng.choice(_CURRENCY_SYMBOLS)
                df[c] = df[c].apply(
                    lambda v: f"{sym}{abs(v):,.2f}"
                    if isinstance(v, (int, float)) and pd.notna(v) else v
                )
                self._report.append(
                    f"  ✗  Currency symbol '{sym}' prefixed on '{c}'."
                )

        # ── 2. Accounting negatives ───────────────────────────────────────────
        if issues.get("accounting_neg", False):
            c = col_for("float")
            if c:
                def _acct(v: Any) -> Any:
                    if isinstance(v, (int, float)) and pd.notna(v):
                        return f"({abs(v):.2f})" if v < 0 else f"{v:.2f}"
                    return v
                df[c] = df[c].apply(_acct)
                self._report.append(
                    f"  ✗  Accounting negatives (n.nn) applied to '{c}'."
                )

        # ── 3. EU decimal format ──────────────────────────────────────────────
        if issues.get("eu_decimal", False):
            c = col_for("float")
            if c:
                def _eu(v: Any) -> Any:
                    if isinstance(v, str):
                        return v            # already formatted
                    if isinstance(v, (int, float)) and pd.notna(v):
                        s = f"{v:,.2f}"     # "1,234.56"
                        return s.replace(",", "X").replace(".", ",").replace("X", ".")
                    return v
                df[c] = df[c].apply(_eu)
                self._report.append(
                    f"  ✗  EU decimal format (1.234,56) applied to '{c}'."
                )

        # ── 4. Percentage suffix ──────────────────────────────────────────────
        if issues.get("percentages", False):
            c = col_for("percentage")
            if c:
                df[c] = df[c].apply(
                    lambda v: f"{v:.1f}%"
                    if isinstance(v, (int, float)) and pd.notna(v) else v
                )
                self._report.append(
                    f"  ✗  Percentage suffix '%' applied to '{c}'."
                )

        # ── 5. Mixed boolean formats ──────────────────────────────────────────
        if issues.get("mixed_booleans", False):
            c = col_for("boolean")
            if c:
                def _mixbool(v: Any) -> str:
                    if v is True or v == "True" or v is True:
                        return self._py_rng.choice(_BOOL_TRUE)
                    return self._py_rng.choice(_BOOL_FALSE)
                df[c] = df[c].apply(_mixbool)
                self._report.append(
                    f"  ✗  Mixed boolean representations in '{c}'."
                )

        # ── 6. Numbers stored as text ─────────────────────────────────────────
        if issues.get("numbers_as_text", False):
            c = col_for("integer")
            if c:
                df[c] = df[c].apply(
                    lambda v: str(int(v))
                    if isinstance(v, (int, float, np.integer)) and pd.notna(v) else v
                )
                self._report.append(
                    f"  ✗  Integers stored as text strings in '{c}'."
                )

        # ── 7. Null variants (~10 % of non-key cells) ─────────────────────────
        if issues.get("null_variants", False):
            skip = {"first_name", "last_name", "email"}
            count = 0
            for ctype, col_name in COLUMN_TYPES.items():
                if ctype in skip:
                    continue
                c = col_for(ctype)
                if c is None or c not in df.columns:
                    continue
                for idx in df.index:
                    if self._py_rng.random() < 0.10:
                        df.at[idx, c] = self._py_rng.choice(_NULL_VARIANTS)
                        count += 1
            self._report.append(
                f"  ✗  Null variants injected into {count} cells (~10 % of value cols)."
            )

        # ── 8. Excel error strings (~2 % of string cells) ─────────────────────
        if issues.get("excel_errors", False):
            str_cols = [c for c in df.columns if df[c].dtype == object]
            count = 0
            for col in str_cols:
                for idx in df.index:
                    if self._py_rng.random() < 0.02:
                        df.at[idx, col] = self._py_rng.choice(_EXCEL_ERRORS)
                        count += 1
            self._report.append(
                f"  ✗  Excel error strings injected into {count} cells (~2 %)."
            )

        # ── 9. Extra whitespace (~8 % of string cells) ────────────────────────
        if issues.get("extra_whitespace", False):
            str_cols = [c for c in df.columns if df[c].dtype == object]
            count = 0
            for col in str_cols:
                for idx in df.index:
                    if pd.notna(df.at[idx, col]) and self._py_rng.random() < 0.08:
                        v      = str(df.at[idx, col])
                        choice = self._py_rng.randint(0, 2)
                        if choice == 0:
                            df.at[idx, col] = "  " + v
                        elif choice == 1:
                            df.at[idx, col] = v + "  "
                        else:
                            df.at[idx, col] = "  " + v + "  "
                        count += 1
            self._report.append(
                f"  ✗  Extra whitespace added to {count} string cells (~8 %)."
            )

        # ── 10. Mixed-case / inconsistent headers ─────────────────────────────
        if issues.get("mixed_case_headers", False):
            rename: dict[str, str] = {}
            styles = ["title", "upper", "lower", "mixed", "original"]
            for c in df.columns:
                style = self._py_rng.choice(styles)
                if style == "upper":
                    rename[c] = c.upper()
                elif style == "lower":
                    rename[c] = c.lower()
                elif style == "mixed":
                    # Alternate caps: "fIrSt NaMe"
                    rename[c] = "".join(
                        ch.upper() if i % 2 == 0 else ch.lower()
                        for i, ch in enumerate(c)
                    )
                else:
                    rename[c] = c
            df = df.rename(columns=rename)
            self._report.append(
                "  ✗  Mixed-case column headers applied."
            )

        # ── 11. Duplicate column header ───────────────────────────────────────
        if issues.get("duplicate_col", False) and len(df.columns) >= 2:
            dup = df.columns[0]
            df.insert(1, dup, df[dup], allow_duplicates=True)
            self._report.append(
                f"  ✗  Column '{dup}' duplicated at position 1."
            )

        # ── 12. Blank columns ─────────────────────────────────────────────────
        if issues.get("blank_cols", False):
            n = max(1, len(df.columns) // 6)
            for i in range(n):
                pos  = self._py_rng.randint(0, len(df.columns))
                name = f"Unnamed: {pos + i}"
                df.insert(min(pos + i, len(df.columns)), name, np.nan)
            self._report.append(
                f"  ✗  {n} fully-empty column(s) inserted."
            )

        # ── 13. Blank rows (~5 %) — must be last structural change ────────────
        if issues.get("blank_rows", False):
            n_blank = max(1, len(df) // 20)
            positions = sorted(
                self._py_rng.sample(range(len(df)), min(n_blank, len(df)))
            )
            blank = pd.DataFrame(
                [[np.nan] * len(df.columns)],
                columns=df.columns,
            )
            pieces: list[pd.DataFrame] = []
            prev = 0
            for pos in positions:
                pieces.append(df.iloc[prev:pos])
                pieces.append(blank)
                prev = pos
            pieces.append(df.iloc[prev:])
            df = pd.concat(pieces, ignore_index=True)
            self._report.append(
                f"  ✗  {n_blank} blank row(s) interspersed."
            )

        return df

    # ── writers ───────────────────────────────────────────────────────────────

    def _write_csv(
        self,
        df:    pd.DataFrame,
        path:  Path,
        bom:   bool        = False,
        title: str | None  = None,
    ) -> None:
        buf = io.StringIO()
        df.to_csv(buf, index=False, lineterminator="\r\n")
        raw = buf.getvalue()

        enc = "utf-8-sig" if bom else "utf-8"
        with open(path, "w", encoding=enc, newline="") as f:
            if title:
                f.write(title + "\r\n")
            f.write(raw)

    def _write_xlsx(
        self,
        df:     pd.DataFrame,
        path:   Path,
        merged: bool       = False,
        hidden: bool       = False,
        title:  str | None = None,
    ) -> None:
        import openpyxl

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Sheet1")
            ws = writer.sheets["Sheet1"]

            if title:
                ws.insert_rows(1)
                ws.cell(row=1, column=1, value=title)
                self._report.append(
                    f"  ✗  Merged cells A2:A4 + title row in Excel."
                )

            if merged and ws.max_row >= 5:
                # Merge first column rows 3–5 (data rows, after optional title+header)
                start = 3 if title else 2
                end   = start + 2
                ws.merge_cells(f"A{start}:A{end}")
                self._report.append(
                    f"  ✗  Cells A{start}:A{end} merged in Excel."
                )

            if hidden and ws.max_row >= 8:
                for r in range(5, 8):
                    ws.row_dimensions[r].hidden = True
                ws.column_dimensions["C"].hidden = True
                self._report.append(
                    "  ✗  Rows 5–7 and column C hidden in Excel."
                )

    # ── helpers ───────────────────────────────────────────────────────────────

    def _find_col(self, df: pd.DataFrame, col_name: str) -> str | None:
        if col_name in df.columns:
            return col_name
        lower_map = {c.lower(): c for c in df.columns}
        return lower_map.get(col_name.lower())
